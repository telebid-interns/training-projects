#!/usr/bin/env python3
import argparse
import collections
import enum
import os
import re
import subprocess
import sys
import time

import openpyxl
import openpyxl.utils

DEFAULT_ROW_WIDTH = 40
DEFAULT_CHAR_WIDTH = 2
COLUMN_ORDER = (
    'Server Software',
    'Server Hostname',
    'Server Port',
    'Document Path',
    'Document Length',
    'Concurrency Level',
    'Time taken for tests',
    'Complete requests',
    'Failed requests',
    'Non-2xx responses',
    'Total transferred',
    'HTML transferred',
    'Requests per second',
    'Time per request',
    'Time per request',
    'Transfer rate'
)
SERVICE_NAME = 'web-server'
WEB_SERVER_DIR = '/opt/web-server/'

TestResult = collections.namedtuple('TestResult', ['commit',
                                                   'cpu_time',
                                                   'mem_usage',
                                                   'worker_time',
                                                   'ab_fields',
                                                   'ab_raw'])
ABResult = collections.namedtuple('ABResult', ['commit', 'fields', 'raw'])
ABField = collections.namedtuple('ABField', ['field_text', 'value_text',
                                             'name', 'value', 'value_suffix'])


class TableExporter:
    def __init__(self, sheet):
        self.inserted_table_positions = []
        self.sheet = sheet
        self.column_widths = {}

    def insert_table(self, table, column_names=None, position=(0, 0),
                     table_name=None, expand_columns=True, expand_rows=True):
        assert isinstance(table, collections.Sequence)
        assert all(isinstance(r, collections.Sequence) for r in table)

        if not table:
            return 0, 0

        row, col = position
        if table_name:
            r, _ = self.insert_row([table_name], (row, col),
                                   expand_row=expand_rows,
                                   expand_columns=False)
            row += r
        if column_names:
            r, _ = self.insert_row(column_names, (row, col),
                                   expand_row=expand_rows,
                                   expand_columns=False)
            row += r
        for row_values in table:
            r, _ = self.insert_row(row_values, (row, col),
                                   expand_row=expand_rows,
                                   expand_columns=expand_columns)
            row += r

        return len(table), len(table[0])

    def insert_row(self, values, position=(0, 0), expand_row=True,
                   expand_columns=True):
        row, col = position
        i = None

        values = tuple(map(str, values))
        for i, v in enumerate(values):
            self.sheet.cell(row=row, column=col + i, value=v)

        if expand_row:
            line_c = max(v.count('\n') for v in values)
            self.sheet.row_dimensions[row].height = DEFAULT_ROW_WIDTH * line_c
        if expand_columns:
            for i, val in enumerate(values):
                letter = openpyxl.utils.get_column_letter(i + 1)
                current_width = self.sheet.column_dimensions[letter].width or 0
                required_width = len(val) * DEFAULT_CHAR_WIDTH
                new_width = max(required_width, current_width)
                self.sheet.column_dimensions[letter].width = new_width
                self.column_widths[letter] = new_width

        if i:
            return 1, i + 1
        else:
            return 0, 0

    def insert_table_from_dicts(self, dicts, column_names, position=(0, 0),
                                table_name=None):
        table = []
        for dct in dicts:
            sorted_ = sorted(dct.items,
                             key=lambda entry: column_names.index(entry[0]))
            table.append(tuple(entry[1] for entry in sorted_))

        return self.insert_table(table=table, column_names=column_names,
                                 position=position, table_name=table_name)


class ABOutTypes(enum.Enum):
    default = 'default'


def parse_default_number(text, value):
    match = re.match(r'([\d.]+)(.*)', value)

    if not match:
        raise ValueError('Value does not contain a number.')

    return ABField(field_text=text, value_text=value,
                   name=text, value=float(match.group(1)),
                   value_suffix=match.group(2))


def parse_default(text, value):
    return ABField(field_text=text, value_text=value,
                   name=text, value=value, value_suffix='')


def parse_ab_out(out, type_=ABOutTypes.default):
    field_regex = re.compile(r'\s*(.+?)\s*:\s*(.*)\s*')

    result = []

    for i, line in enumerate(out.splitlines()):
        line = line.strip()
        if i < 7 or not line:
            continue

        match = field_regex.match(line)

        if not match:
            continue

        field_text, value_text = match.group(1), match.group(2)

        if field_text not in COLUMN_ORDER:
            continue

        try:
            ab_field = parse_default_number(field_text, value_text)
        except ValueError:
            ab_field = parse_default(field_text, value_text)

        result.append(ab_field)

    return result


def run_test(ab_args):
    ab_args = tuple(str(a) for a in ab_args)
    print('Restarting {} service'.format(SERVICE_NAME))
    subprocess.check_output(['systemctl', 'restart', SERVICE_NAME])
    # sleep some time for the server to restart successfully or AB will break
    time.sleep(2)

    print('Running AB with arguments: `{}`'
          .format(' '.join(ab_args)))
    out = subprocess.check_output(ab_args).decode('utf-8')
    fields = parse_ab_out(out)

    git_out = subprocess.check_output(['git', 'log', '--oneline'],
                                      universal_newlines=True)
    commit = git_out.splitlines()[0]

    rusage_path = os.path.join(WEB_SERVER_DIR, 'bin/rusage.sh')
    access_log_path = os.path.join(WEB_SERVER_DIR, 'data/logs/access.log')
    try:
        rusage_out = subprocess.check_output([rusage_path, access_log_path])
        cpu_time, mem_usage, worker_time, *_ = (rusage_out.decode('ascii')
                                                .split(' '))
    except (OSError, ValueError, subprocess.CalledProcessError) as e:
        print(e)
        cpu_time, mem_usage, worker_time = ('-', '-', '-')

    return TestResult(commit=commit, ab_fields=fields, ab_raw=out,
                      cpu_time=cpu_time, mem_usage=mem_usage,
                      worker_time=worker_time)


def run_tests(host, port, *, request_numbers, concurrences, static_route,
              benchmarks_dir, protocol='http'):
    os.makedirs(benchmarks_dir, exist_ok=True)
    if static_route.startswith('/'):
        static_route = static_route[1:]

    results = []
    for n, c in zip(request_numbers, concurrences):
        url = '{}://{}:{}/{}'.format(protocol, host, port, static_route)
        ab_args = ['ab', '-n', n, '-c', c, url]
        ab_result = run_test(ab_args)

        test_name = "GET_{}_{}_{}".format(
            static_route.replace('/', percent_encode('/')),
            n,
            c
        )
        raw_file_name = os.path.join(
            benchmarks_dir, test_name + '.raw'
        )
        with open(raw_file_name, mode='w') as f:
            f.write(ab_result.ab_raw)

        # json_file_name = os.path.join(benchmarks_dir, test_name + '.json')
        # with open(json_file_name, mode='w') as f:
        #     json.dump(ab_result_to_dict(ab_result), f)

        results.append(ab_result)

    xlsx_file_name = os.path.join(
        benchmarks_dir,
        "GET_{}.xlsx".format(static_route).replace('/', percent_encode('/'))
    )
    dump_test_results_to_xlsx(
        results,
        xlsx_file_name
    )


def dump_test_results_to_xlsx(test_results, file_path):
    assert len(set(ar.ab_fields[0] for ar in test_results)) == 1

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Benchmarks.'
    row, col = (1, 1)
    te = TableExporter(sheet)
    table = []
    column_names = None

    for test_result in test_results:
        extra = (test_result.cpu_time, test_result.mem_usage,
                 test_result.worker_time)
        extra = list(map(str, extra))
        fields = test_result.ab_fields

        if not column_names:
            column_names = [
                '{}\n({})'.format(field.field_text, field.value_suffix)
                for field in fields]
            column_names.extend(['cpu_time', 'mem_usage', 'worker_time'])

        table_row = [field.value for field in fields] + extra
        table.append(table_row)

    te.insert_table(table=table,
                    column_names=column_names,
                    position=(row, col), table_name='Apache Benchmark results',
                    expand_columns=True, expand_rows=True)

    wb.save(file_path)

    return


def namedtuple_to_dict(nt):
    # noinspection PyProtectedMember
    dct = nt._asdict()

    for key, value in dct.items():
        if hasattr(value, '_asdict'):
            dct[key] = namedtuple_to_dict(value)

    return dct


def percent_encode(char):
    return '%' + hex(ord(char))[2:]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'concurrences',
        help='Space separated list of concurrences to benchmark',
        nargs='+',
        type=int
    )
    parser.add_argument('--host', help='Host name of server.',
                        default='localhost')
    parser.add_argument('-p', '--port', help='Port of server.',
                        default='5678', type=int)
    parser.add_argument('-r', help='Route to benchmark', required=True)
    parser.add_argument('-n', help='Number of total requests', default=20)
    parser.add_argument('-d', '--directory', default='tests/benchmarks/',
                        help='Output directory for results.')
    args = parser.parse_args()

    if not os.path.exists(args.directory):
        os.makedirs(args.directory)
    elif not os.path.isdir(args.directory):
        print('{} is not a directory'.format(args.directory))
        sys.exit(1)

    request_numbers = [args.n] * len(args.concurrences)
    run_tests(args.host, args.port,
              request_numbers=request_numbers,
              concurrences=args.concurrences,
              static_route=args.r,
              benchmarks_dir=args.directory)


if __name__ == '__main__':
    main()
